# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
import openpyxl
from pdf2image import convert_from_path
import os
import glob


   

# Load in the workbook
#wb = load_workbook('./Копия Свод ТЕ для РКМ.xlsx')
#sheet = wb.active
for i in range (1, 145):
#     #for g in range (1, 145):
#         #if  sheet ['A'+ str (i)].value.split('ЛЦУИ')[-1] == sheet ['A'+ str (g)].value.split('ЛЦУИ')[-1]:
#            # print(sheet ['A'+ str (i)].value, sheet ['A'+ str (g)].value)
#             #if i != g:
#                 #print(i)
#    sheet['B' + str(i)].hyperlink = "http://stackoverflow.com"   #гипперссылка на чертеж                                              

#wb.save('test.xlsx')




for file in glob.glob('555-03.087.*'):
    print(file)
    
